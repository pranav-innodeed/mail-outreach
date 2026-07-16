"""Excel-driven Outlook Web outreach through Selenium.

Start in dry-run mode. Drafts are created but not sent until config.json has
"dry_run": false. This program deliberately requires interactive sign-in.
"""

from __future__ import annotations

import json
import logging
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import SessionNotCreatedException, TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"
LOG_PATH = BASE_DIR / "outreach.log"

RECIPIENT_HEADERS = ["Name", "Company", "Email", "Template Key", "Status", "Sent At", "Notes"]
TEMPLATE_HEADERS = ["Template Key", "Subject", "Body"]
READY_STATUSES = {"", "READY", "FAILED"}


class OutreachError(RuntimeError):
    """Expected problem that should be recorded against a recipient."""


@dataclass(frozen=True)
class Recipient:
    row_number: int
    name: str
    company: str
    email: str
    template_key: str


@dataclass(frozen=True)
class Template:
    subject: str
    body: str


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[logging.FileHandler(LOG_PATH, encoding="utf-8"), logging.StreamHandler()],
    )


def load_config() -> dict[str, Any]:
    try:
        with CONFIG_PATH.open(encoding="utf-8") as file:
            config = json.load(file)
    except json.JSONDecodeError as error:
        raise OutreachError(
            "config.json is not valid JSON. For Windows paths, use forward slashes "
            '(C:/Users/name/file.xlsx) or escape each backslash (C:\\\\Users\\\\name\\\\file.xlsx).'
        ) from error
    required = {"workbook_path", "outlook_url", "chrome_profile_path", "dry_run", "max_emails_per_run", "delay_between_emails_seconds", "wait_timeout_seconds"}
    missing = required - config.keys()
    if missing:
        raise OutreachError(f"config.json is missing: {', '.join(sorted(missing))}")
    return config


def header_map(sheet, expected: list[str]) -> dict[str, int]:
    actual = [sheet.cell(1, col).value for col in range(1, len(expected) + 1)]
    if actual != expected:
        raise OutreachError(f"Sheet '{sheet.title}' must use headers: {expected}")
    return {name: index + 1 for index, name in enumerate(expected)}


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def read_templates(sheet) -> dict[str, Template]:
    columns = header_map(sheet, TEMPLATE_HEADERS)
    templates: dict[str, Template] = {}
    for row in range(2, sheet.max_row + 1):
        key = text(sheet.cell(row, columns["Template Key"]).value).lower()
        subject = text(sheet.cell(row, columns["Subject"]).value)
        body = text(sheet.cell(row, columns["Body"]).value)
        if not any((key, subject, body)):
            continue
        if not all((key, subject, body)):
            raise OutreachError(f"Templates row {row} needs a key, subject, and body.")
        if key in templates:
            raise OutreachError(f"Template key '{key}' is duplicated.")
        templates[key] = Template(subject=subject, body=body)
    if not templates:
        raise OutreachError("No templates were found.")
    return templates


def read_ready_recipients(sheet, templates: dict[str, Template]) -> tuple[list[Recipient], dict[str, int]]:
    columns = header_map(sheet, RECIPIENT_HEADERS)
    recipients: list[Recipient] = []
    for row in range(2, sheet.max_row + 1):
        name = text(sheet.cell(row, columns["Name"]).value)
        company = text(sheet.cell(row, columns["Company"]).value)
        email = text(sheet.cell(row, columns["Email"]).value)
        template_key = text(sheet.cell(row, columns["Template Key"]).value).lower()
        status = text(sheet.cell(row, columns["Status"]).value).upper()
        if not any((name, company, email, template_key, status)):
            continue
        if status == "SENT":
            continue
        if status not in READY_STATUSES:
            logging.info("Skipping row %s with status %r", row, status)
            continue
        if not name or not company or not email or not template_key:
            raise OutreachError(
                f"Recipients row {row} needs Name, Company, Email, and Template Key."
            )
        if not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", email):
            raise OutreachError(f"Recipients row {row} has an invalid email address: {email}")
        if template_key not in templates:
            raise OutreachError(f"Recipients row {row} uses unknown template key '{template_key}'.")
        recipients.append(Recipient(row, name, company, email, template_key))
    return recipients, columns


def first_name(full_name: str) -> str:
    """Use only the first word of the recipient name for greetings."""
    parts = full_name.split()
    return parts[0] if parts else full_name


def personalise(value: str, recipient: Recipient) -> str:
    first = first_name(recipient.name)
    return (
        value.replace("{{First Name}}", first)
        .replace("{{first name}}", first)
        .replace("{{name}}", first)
        .replace("{{Company Name}}", recipient.company)
        .replace("{{company}}", recipient.company)
    )


def stop_profile_chrome_instances(profile_path: Path) -> int:
    """Stop orphaned Selenium Chrome processes that still hold the profile open."""
    profile = str(profile_path.resolve()).lower()
    if sys.platform == "win32":
        command = (
            "Get-CimInstance Win32_Process -Filter \"name='chrome.exe'\" | "
            "Where-Object { "
            f"$_.CommandLine -and $_.CommandLine.ToLower().Contains('{profile}') "
            "-and $_.CommandLine -like '*enable-automation*' "
            "} | ForEach-Object { Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }"
        )
        subprocess.run(
            ["powershell", "-NoProfile", "-Command", command],
            check=False,
            capture_output=True,
            text=True,
        )
        return 0
    escaped = re.escape(profile)
    result = subprocess.run(
        ["pgrep", "-f", escaped],
        check=False,
        capture_output=True,
        text=True,
    )
    stopped = 0
    for pid in result.stdout.split():
        if pid.isdigit():
            subprocess.run(["kill", "-9", pid], check=False)
            stopped += 1
    return stopped


def clear_stale_profile_locks(profile_path: Path) -> None:
    """Remove leftover lock files after a crashed or interrupted browser session."""
    for name in ("lockfile", "SingletonLock", "SingletonCookie", "SingletonSocket"):
        (profile_path / name).unlink(missing_ok=True)
    (profile_path / "Default" / "LOCK").unlink(missing_ok=True)


def start_browser(profile_path: Path) -> webdriver.Chrome:
    profile_path.mkdir(parents=True, exist_ok=True)
    stop_profile_chrome_instances(profile_path)
    clear_stale_profile_locks(profile_path)
    options = webdriver.ChromeOptions()
    options.add_argument(f"--user-data-dir={profile_path}")
    options.add_argument("--profile-directory=Default")
    options.add_argument("--start-maximized")
    try:
        return webdriver.Chrome(options=options)
    except SessionNotCreatedException as error:
        raise OutreachError(
            "Chrome could not start with the saved profile. Close any Chrome windows opened by "
            "outreach.py, delete the folder in chrome_profile_path if the problem continues, "
            "then run the script again."
        ) from error


def first_visible(wait: WebDriverWait, locators: list[tuple[str, str]]):
    def locate(driver):
        for locator in locators:
            for element in driver.find_elements(*locator):
                if element.is_displayed():
                    return element
        return False
    return wait.until(locate)


OUTLOOK = {
    "new_message": [
        (By.CSS_SELECTOR, "button[aria-label*='New mail'], button[aria-label*='New message'], button[aria-label*='New email'], button[title*='New mail']"),
        (By.CSS_SELECTOR, "[role='button'][aria-label*='New mail'], [role='button'][aria-label*='New message'], [role='button'][aria-label*='New email']"),
        (By.XPATH, "//*[self::button or @role='button'][normalize-space()='New mail' or normalize-space()='New message' or normalize-space()='New email' or normalize-space()='Compose' or .//*[normalize-space()='New mail' or normalize-space()='New message' or normalize-space()='New email' or normalize-space()='Compose']]")],
    "to": [
        (By.XPATH, "//input[@aria-label='To' or @placeholder='To']"),
        (By.XPATH, "//*[(@role='textbox' or @contenteditable='true') and (contains(@aria-label, 'To') or contains(@data-automation-id, 'to'))]"),
    ],
    "from_control": [(By.XPATH, "//*[self::button or @role='button'][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'from:')]")],
    "to_label": [(By.XPATH, "//*[self::button or @role='button'][normalize-space()='To']")],
    "subject": [
        (By.XPATH, "//input[contains(@aria-label, 'Subject') or contains(@placeholder, 'Subject')]"),
        (By.XPATH, "//*[(@role='textbox' or @contenteditable='true') and (contains(translate(@aria-label, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'subject') or contains(translate(@placeholder, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'subject'))]"),
    ],
    "subject_label": [(By.XPATH, "//*[normalize-space()='Add a subject']")],
    "body": [
        (By.XPATH, "//*[@contenteditable='true' and (contains(translate(@aria-label, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'message') or contains(translate(@aria-label, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'body'))]"),
        (By.XPATH, "//*[@role='textbox' and (contains(translate(@aria-label, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'message') or contains(translate(@aria-label, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'body'))]"),
    ],
    "send": [(By.XPATH, "//button[@aria-label='Send' or .//*[normalize-space()='Send']]")],
    "error_dialog": [(By.XPATH, "//*[@role='alertdialog' or @role='alert']")],
}


def fill_message_body(driver: webdriver.Chrome, body_field, message: str) -> None:
    """Insert body text without ChromeDriver send_keys (emoji / non-BMP chars fail there)."""
    focus_element(driver, body_field)
    driver.execute_script(
        """
        const editor = arguments[0];
        const text = arguments[1];
        editor.focus();
        const selection = window.getSelection();
        const range = document.createRange();
        range.selectNodeContents(editor);
        selection.removeAllRanges();
        selection.addRange(range);
        const inserted = document.execCommand('insertText', false, text);
        if (!inserted || !(editor.innerText || editor.textContent || '').trim()) {
            editor.innerText = text;
            editor.dispatchEvent(new InputEvent('input', { bubbles: true, data: text }));
        }
        """,
        body_field,
        message,
    )


def compose_and_optionally_send(driver: webdriver.Chrome, config: dict[str, Any], recipient: Recipient, template: Template) -> str:
    wait = WebDriverWait(driver, int(config["wait_timeout_seconds"]))
    new_message = first_visible(wait, OUTLOOK["new_message"])
    wait.until(EC.element_to_be_clickable(new_message)).click()
    select_sender(driver, wait, text(config.get("from_email")))
    to_field = find_recipient_field(driver, wait)
    to_field.send_keys(recipient.email)
    to_field.send_keys(Keys.ENTER)
    subject_field = find_subject_field(driver, wait)
    subject_field.send_keys(personalise(template.subject, recipient))
    body_field = find_message_body(driver, wait, subject_field)
    fill_message_body(driver, body_field, personalise(template.body, recipient))
    if config["dry_run"]:
        return "DRAFT CREATED (dry run)"
    send_button = first_visible(wait, OUTLOOK["send"])
    wait.until(EC.element_to_be_clickable(send_button)).click()
    time.sleep(2)
    for locator in OUTLOOK["error_dialog"]:
        for alert in driver.find_elements(*locator):
            if alert.is_displayed() and text(alert.text):
                raise OutreachError(f"Outlook reported: {text(alert.text)[:200]}")
    return "SENT"


def select_sender(driver: webdriver.Chrome, wait: WebDriverWait, from_email: str) -> None:
    """Select an Outlook sender that is already available to the signed-in user."""
    if not from_email:
        return
    from_control = first_visible(wait, OUTLOOK["from_control"])
    from_control.click()
    email_lower = from_email.lower()

    def matching_sender(browser):
        candidates = browser.find_elements(
            By.XPATH,
            "//*[@role='menuitem' or @role='option' or @role='menuitemradio']",
        )
        for candidate in candidates:
            if candidate.is_displayed() and email_lower in candidate.text.lower():
                return candidate
        return False

    try:
        sender = wait.until(matching_sender)
    except TimeoutException as error:
        raise OutreachError(
            f"'{from_email}' is not available in Outlook's From menu. "
            "Use an account with Send As/Send on Behalf permission, or leave from_email blank."
        ) from error
    sender.click()


def focused_text_field(driver: webdriver.Chrome):
    element = driver.switch_to.active_element
    if not element.is_displayed():
        return False
    if element.tag_name in {"input", "textarea"} or element.get_attribute("contenteditable") == "true":
        return element
    return False


def find_recipient_field(driver: webdriver.Chrome, wait: WebDriverWait):
    """Support both classic input fields and the current Outlook people picker."""
    short_wait = WebDriverWait(driver, 3)
    try:
        field = first_visible(short_wait, OUTLOOK["to"])
        field.click()
        return field
    except TimeoutException:
        to_label = first_visible(wait, OUTLOOK["to_label"])
        to_label.click()
        return wait.until(focused_text_field)


def find_subject_field(driver: webdriver.Chrome, wait: WebDriverWait):
    """Support both an input-based subject field and Outlook's custom editor."""
    short_wait = WebDriverWait(driver, 3)
    try:
        field = first_visible(short_wait, OUTLOOK["subject"])
        focus_element(driver, field)
        return field
    except TimeoutException:
        subject_label = first_visible(wait, OUTLOOK["subject_label"])
        driver.execute_script("arguments[0].click();", subject_label)
        return wait.until(focused_text_field)


def focus_element(driver: webdriver.Chrome, element) -> None:
    """Focus an editor without clicking through Outlook's recipient suggestion layer."""
    driver.execute_script(
        "arguments[0].scrollIntoView({block: 'center', inline: 'nearest'}); arguments[0].focus();",
        element,
    )


def find_message_body(driver: webdriver.Chrome, wait: WebDriverWait, subject_field):
    """Find the message editor without accidentally selecting To/Cc people pickers."""
    short_wait = WebDriverWait(driver, 3)
    try:
        return first_visible(short_wait, OUTLOOK["body"])
    except TimeoutException:
        subject_y = subject_field.rect["y"]

        def message_editor_below_subject(browser):
            candidates = browser.find_elements(By.XPATH, "//*[@contenteditable='true' or @role='textbox']")
            valid = []
            for candidate in candidates:
                if not candidate.is_displayed():
                    continue
                label = " ".join([
                    candidate.get_attribute("aria-label") or "",
                    candidate.get_attribute("placeholder") or "",
                ]).lower()
                rect = candidate.rect
                is_recipient_field = any(word in label for word in ("to", "cc", "bcc", "recipient"))
                if not is_recipient_field and rect["y"] > subject_y and rect["width"] > 250:
                    valid.append(candidate)
            return valid[0] if valid else False

        return wait.until(message_editor_below_subject)


def save_result(workbook, sheet, row: int, columns: dict[str, int], status: str, note: str, workbook_path: Path) -> bool:
    sheet.cell(row, columns["Status"]).value = status
    sheet.cell(row, columns["Sent At"]).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.cell(row, columns["Notes"]).value = note
    try:
        workbook.save(workbook_path)
        return True
    except PermissionError:
        logging.error(
            "Could not update %s because it is open or locked. Close the workbook in Excel, "
            "then run the script again.", workbook_path,
        )
        return False


def capture_ui_failure(driver: webdriver.Chrome) -> Path | None:
    """Save a local screenshot to make Outlook selector issues diagnosable."""
    screenshot = BASE_DIR / f"outlook_failure_{datetime.now():%Y%m%d_%H%M%S}.png"
    try:
        driver.save_screenshot(str(screenshot))
        logging.error("Saved Outlook diagnostic screenshot: %s", screenshot)
        return screenshot
    except WebDriverException:
        logging.exception("Could not save Outlook diagnostic screenshot.")
        return None


def main() -> int:
    configure_logging()
    config = load_config()
    workbook_path = BASE_DIR / config["workbook_path"]
    if not workbook_path.exists():
        raise OutreachError(f"Workbook not found: {workbook_path}.")
    workbook = load_workbook(workbook_path)
    if "Recipients" not in workbook.sheetnames or "Templates" not in workbook.sheetnames:
        raise OutreachError("Workbook must contain 'Recipients' and 'Templates' sheets.")
    recipients_sheet = workbook["Recipients"]
    templates = read_templates(workbook["Templates"])
    recipients, columns = read_ready_recipients(recipients_sheet, templates)
    recipients = recipients[:max(0, int(config["max_emails_per_run"]))]
    if not recipients:
        logging.info("No eligible recipients found.")
        return 0
    driver = start_browser((BASE_DIR / config["chrome_profile_path"]).resolve())
    try:
        driver.get(config["outlook_url"])
        input("Sign in to Outlook in Chrome, wait for your inbox, then press Enter here... ")
        for number, recipient in enumerate(recipients, start=1):
            try:
                result = compose_and_optionally_send(driver, config, recipient, templates[recipient.template_key])
                status = "DRAFT" if config["dry_run"] else "SENT"
                save_result(workbook, recipients_sheet, recipient.row_number, columns, status, result, workbook_path)
                logging.info("%s: %s", status, recipient.email)
            except (OutreachError, TimeoutException, WebDriverException) as error:
                if isinstance(error, TimeoutException):
                    screenshot = capture_ui_failure(driver)
                    message = "Timed out waiting for an Outlook control."
                    if screenshot:
                        message += f" Screenshot: {screenshot.name}"
                else:
                    message = str(error) or type(error).__name__
                message = message[:250]
                save_result(workbook, recipients_sheet, recipient.row_number, columns, "FAILED", message, workbook_path)
                logging.exception("Failed for %s", recipient.email)
            if number < len(recipients):
                time.sleep(max(0, float(config["delay_between_emails_seconds"])))
    finally:
        input("Review the browser window. Press Enter to close it... ")
        driver.quit()
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except OutreachError as error:
        logging.error("%s", error)
        raise SystemExit(1)
